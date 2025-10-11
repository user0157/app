import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pages.page import Page  # 自定义的页面类，用于处理页面切换
from widgets import Mytable, MyText, ToastMessage  # 自定义的控件：表格、文本编辑器、提示信息
from pages.utils import wrap_number_to_lines, generate_product_map, load_cache, generate_aliases  # 工具函数
from openpyxl import Workbook, load_workbook  # 用于处理Excel文件
from openpyxl.utils import get_column_letter  # 计算Excel中的列字母
from openpyxl.styles import Font  # 用于Excel中的字体样式
import datetime
import pyperclip  # 用于操作剪贴板


class MainPage(Page):
    def __init__(self, parent, page_id):
        """
        初始化主页面。
        :param parent: 父级控件
        :param page_id: 页面唯一标识
        """
        super().__init__(parent, page_id)  # 调用父类（Page）的构造函数


    def create_widgets(self):
        """
        创建并排列主页面中的控件。
        包括侧边栏、文本编辑器和表格。
        """
        # 配置网格布局，使其具有弹性伸缩性
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=0)  # 侧边栏列
        self.grid_columnconfigure(1, weight=1)  # 主内容列

        # 侧边栏框架配置
        self.sidebar = ttk.Frame(self, width=150)
        self.sidebar.grid(row=0, column=0, sticky="ns", padx=5, pady=5)
        self.sidebar.grid_propagate(False)  # 防止侧边栏大小随内容变化

        self.sidebar.grid_columnconfigure(0, weight=1)  # 侧边栏按钮占满整个宽度

        # 按钮列表，每个按钮包含文字和对应的动作（命令）
        buttons = [
            ("文本标记", None),
            ("标记数量-前面", lambda: self.run_algorithm("qty", "forward")),
            ("标记数量-后面", lambda: self.run_algorithm("qty", "backward")),
            ("标记价格-前面", lambda: self.run_algorithm("price", "forward")),
            ("标记价格-后面", lambda: self.run_algorithm("price", "backward")),
            ("文本处理", None),
            ("匹配文本内容", self.process_text),
            ("表格处理", None),
            ("匹配表格内容", self.get_sku),
            ("获取价格", self.get_price),
            ("添加行", lambda: self.table.add_row()),  # 在表格中添加新行
            ("清理表格数据", lambda: self.table.set_data([])),  # 清空表格数据
            ("复制表格数据", self.copy_to_clipboard),  # 将表格数据复制到剪贴板
            ("Excel", None),
            ("导入Excel", self.upload_excel),  # 导入Excel文件中的数据
            ("导出Excel", self.export_excel),  # 将数据导出为Excel文件
            ("创建缓存", self.create_cache),  # 创建缓存
        ]
        
        # 动态创建按钮或标签，基于按钮列表的内容
        for i, (text, command) in enumerate(buttons):
            if command is None:
                # 如果没有命令，创建标签
                label = ttk.Label(self.sidebar, text=text, anchor="center", font=("Arial", 10, "bold"))
                label.grid(row=i, column=0, sticky="ew", pady=(10, 2))  # 标签居中，并有一些垂直间距
            else:
                # 否则创建按钮，并绑定对应的命令
                btn = ttk.Button(self.sidebar, text=text, command=command)
                btn.grid(row=i, column=0, sticky="ew", pady=2)  # 按钮占满宽度

        # 确保最后一行（空行）可以填充剩余空间
        self.sidebar.grid_rowconfigure(len(buttons), weight=1)

        # 主内容区容器，包含文本编辑器和表格
        self.container = ttk.Frame(self)
        self.container.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        
        # 配置容器内的网格布局，确保它们有弹性
        self.container.grid_rowconfigure(0, weight=1)
        self.container.grid_columnconfigure(0, weight=2, uniform="equal")  # 文本编辑器占更大空间
        self.container.grid_columnconfigure(1, weight=3, uniform="equal")  # 表格占更大空间

        # 文本编辑器，用于文本处理
        self.editor = MyText(self.container)
        self.editor.grid(row=0, column=0, sticky="nsew", padx=(0, 5))  # 编辑器占据左半部分

        # 表格控件，用于显示和操作数据
        self.table = Mytable(self.container)
        self.table.grid(row=0, column=1, sticky="nsew")  # 表格占据右半部分

        # 设置表格的列（sku, name, quantity, price）
        self.table.set_columns(["sku", "name", "quantity", "price"])

    def copy_to_clipboard(self):
        """
        将表格中的数据复制到剪贴板。
        从表格中提取每行的数量、名称和价格，拼接为文本格式后复制到剪贴板。
        """
        item_ids = self.table.tree.get_children()  # 获取表格中所有行的ID
        col_names = self.table.columns  # 获取表格的列名

        lines = []  # 用于存储每行拼接后的文本

        # 遍历每一行，将数量、名称和价格提取并拼接
        for item_id in item_ids:
            values = self.table.tree.item(item_id)['values']  # 获取行的数据
            row_dict = dict(zip(col_names, values))  # 将列名与行数据组合成字典

            # 获取每个字段的值（如果没有值则使用空字符串）
            quantity = row_dict.get('quantity', '')
            name = row_dict.get('name', '')
            price = row_dict.get('price', '')

            # 拼接行文本
            line = f"{quantity} {name} {price}".strip()  # 清理多余的空格
            lines.append(line)

        # 将所有行的数据合并为一个字符串，并复制到剪贴板
        text = "\n".join(lines)
        pyperclip.copy(text)  # 使用 pyperclip 库复制文本
        ToastMessage(self, "成功复制到粘贴面板")  # 显示提示消息

    def logout(self):
        """
        处理登出操作，弹出确认对话框，确认后触发页面退出事件。
        """
        if messagebox.askyesno("Confirm Logout", "Are you sure you want to logout?"):  # 弹出确认框
            self.emit("page/logout", self.page_id)  # 确认登出后发出登出信号

    def run_algorithm(self, key, direction):
        """
        根据选择的标记类型和方向运行算法（如标记数量、标记价格）。
        """
        text = self.editor.get("1.0", tk.END)  # 获取文本框中的内容
        new_text = wrap_number_to_lines(text, key, direction)  # 使用 wrap_number_to_lines 包装文本
        self.editor.overwrite(new_text)  # 用新的文本覆盖原有文本

    def get_sku(self):
        """
        获取表格中的 SKU 数据，将其整理为特定格式的文本，并触发页面处理操作。
        """
        names = [item["name"] for item in self.table.get_data()]  # 获取所有商品名称
        quantitys = [item["quantity"] for item in self.table.get_data()]  # 获取所有数量
        prices = [item["price"] for item in self.table.get_data()]  # 获取所有价格
        
        # 将表格数据格式化为 [qty:数量] 名称 [price:价格] 的格式
        text = "\n".join(f"[qty:{q}] {n} [price:{p}]" for n, q, p in zip(names, quantitys, prices) if n.strip())
        
        # 触发页面处理操作，将整理好的文本传递给页面处理
        self.emit("page/process", text, self.page_id)

    def process_text(self):
        """
        处理文本框中的内容，将其传递给页面处理操作。
        """
        text = self.editor.get("1.0", tk.END)  # 获取文本框中的内容
        self.emit("page/process", text, self.page_id)  # 触发页面处理操作

    def handle_text_processed(self, data):
        """
        处理从页面处理返回的数据，过滤出所需的字段并更新表格。
        """
        desired_keys = ["sku", "name", "quantity", "price"]  # 需要提取的字段

        # 过滤出每一项的这些字段
        filtered = [
            {key: d[key] for key in desired_keys}
            for d in data
        ]

        # 根据 SKU 匹配数据
        result = self.match(filtered, mode="sku")
        if not result:
            return  # 如果没有匹配结果，则不做任何操作

        # 将匹配到的结果更新到表格中
        self.table.set_data(result)  # 更新表格数据
        self.table.set_columns(["sku", "name", "quantity", "price"])  # 设置表格的列名
        self.table.set_editable_columns(["name", "quantity", "price"])  # 设置可编辑的列


    def match(self, data, mode="both"):
        """
        将输入数据与缓存中的产品数据进行匹配。
        根据指定的模式（"sku", "price", 或 "both"）来更新输入数据中的 sku 和 price 字段。
        """
        cache = load_cache()  # 加载缓存
        if not cache:
            return  # 如果缓存为空，则返回空

        price_map = generate_product_map(cache)  # 生成商品的 SKU 到产品的映射

        # 遍历数据中的每一项（如每个商品）
        for item in data:
            original_sku = item.get("sku")  # 获取原始的 SKU
            aliases = generate_aliases(original_sku)  # 获取 SKU 的别名列表

            matched_price = 0  # 初始化匹配的价格为 0
            matched_product = None  # 初始化匹配的产品为 None

            # 遍历每个别名，查找对应的产品
            for alias in aliases:
                product = price_map.get(alias)  # 在 price_map 中查找该别名对应的产品
                if product:
                    price = product.get("price", 0)  # 获取产品的价格
                    if price > matched_price:  # 如果找到更高的价格，更新匹配的价格和产品
                        matched_price = price
                        matched_product = product

            # 根据模式更新数据中的价格和 SKU
            if mode in ("price", "both"):
                item["price"] = "" if matched_price == 0 else matched_price  # 如果没有匹配到价格，则留空

            if mode in ("sku", "both"):
                if matched_product:
                    item["sku"] = matched_product.get("sku", "")  # 如果找到了匹配的产品，更新 SKU
                else:
                    item["sku"] = ""  # 如果没有匹配到产品，SKU 设为空

        return data  # 返回更新后的数据

    def get_price(self):
        """
        获取表格数据中的价格信息，并根据缓存更新每个条目的价格。
        """
        data = self.table.get_data()  # 获取表格中的所有数据
        result = self.match(data, mode="price")  # 通过 match 方法匹配价格
        if not result:
            return  # 如果没有结果，则返回

        self.table.set_data(result)  # 更新表格中的数据

    def create_cache(self):
        """
        创建缓存，触发页面的缓存创建操作。
        """
        self.emit("page/create_cache", self.page_id)  # 触发页面的创建缓存事件
        
    def upload_excel(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")],
            title="选择要导入的Excel文件"
        )
        if not file_path:
            return  # 用户未选择文件时退出

        try:
            wb = load_workbook(file_path)  # 打开Excel文件
            ws = wb.active  # 获取当前工作表

            # 获取表头
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]  
            expected_columns = ["sku", "name", "quantity", "price"]
            existing_columns = [col for col in expected_columns if col in headers]

            if not existing_columns:
                messagebox.showerror("错误", "Excel 文件中没有找到可导入的列。")
                return  # 如果没有找到预期列则提示错误并返回

            col_indices = {col: headers.index(col) for col in existing_columns}  # 获取每列的索引

            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):  # 从第二行开始读取数据
                if all(cell is None for cell in row):  # 如果整行都为空则跳过
                    continue

                item = {}
                for col in existing_columns:
                    value = row[col_indices[col]]

                    # 对 quantity 列进行类型转换，确保是整数
                    if col == "quantity":
                        try:
                            value = int(float(value))
                        except (TypeError, ValueError):
                            value = 0

                    # 对 price 列进行类型转换，确保是浮动的数值，若是整数，则显示为整数
                    elif col == "price":
                        try:
                            price_val = float(value)
                            if price_val.is_integer():
                                value = str(int(price_val))
                            else:
                                value = f"{price_val:.2f}"
                        except (TypeError, ValueError):
                            value = "0"

                    # 如果某个字段为空，填充默认值
                    if value is None:
                        value = "" if col != "quantity" else 0

                    item[col] = value

                # 填充未在 Excel 中出现的列
                for col in expected_columns:
                    if col not in existing_columns:
                        item[col] = 0 if col == "quantity" else ""

                data.append(item)

            # 更新表格数据
            self.table.set_columns(expected_columns)
            self.table.set_data(data)
            self.table.set_editable_columns(expected_columns)
            messagebox.showinfo("成功", "Excel 文件导入成功。")

        except Exception as e:
            messagebox.showerror("错误", f"导入 Excel 文件失败: {str(e)}")

    def export_excel(self):
        default_filename = f"data_{datetime.date.today()}.xlsx"  # 默认文件名

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",  # 默认扩展名
            filetypes=[("Excel files", "*.xlsx *.xls")],  # 允许选择的文件类型
            initialfile=default_filename  # 初始文件名
        )
        
        if not file_path:
            return  # 用户未选择保存文件时退出

        try:
            treeview = self.table.tree  # 获取表格的数据
            wb = Workbook()  # 创建新的工作簿
            ws = wb.active  # 获取活动工作表

            columns = self.table.columns  # 获取表格的列名
            header_font = Font(name='Calibri', size=11, bold=True)  # 设置表头字体
            default_font = Font(name='Calibri', size=11)  # 设置默认字体

            # 写入表头
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font

            # 写入数据
            for row_idx, row_id in enumerate(treeview.get_children(), start=2):
                row_values = treeview.item(row_id)["values"]
                for col_idx, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = default_font

            # 自动调整列宽
            for col_idx, col_name in enumerate(columns, start=1):
                max_length = len(str(col_name))
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = adjusted_width

            # 设置字体
            max_row = ws.max_row
            max_col = ws.max_column
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    if cell.font != header_font:
                        cell.font = default_font

            # 保存 Excel 文件
            wb.save(file_path)
            messagebox.showinfo("成功", "数据导出成功。")
        except Exception as e:
            messagebox.showerror("错误", f"保存 Excel 文件失败: {str(e)}")
